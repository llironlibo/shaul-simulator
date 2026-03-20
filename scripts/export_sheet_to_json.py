#!/usr/bin/env python3
"""
Export Google Sheet data to questions.json for SHAUL Simulator.

Usage (from project root):
  python scripts/export_sheet_to_json.py --xlsx sheet.xlsx --output src/data/questions.json
  python scripts/export_sheet_to_json.py --statements statements.csv --pairs pairs.csv --output src/data/questions.json

Google Sheet structure:
  Sheet 1 "Statements": columns id, text_he, trait, notes
  Sheet 2 "Pairs": columns id, statement_a, statement_b, notes

Valid traits: Conscientiousness, Agreeableness, EmotionalStability, Extraversion, Openness
"""

import argparse
import csv
import json
import os
import sys
from collections import Counter

VALID_TRAITS = {
    "Conscientiousness",
    "Agreeableness",
    "EmotionalStability",
    "Extraversion",
    "Openness",
}


def read_xlsx(path):
    """Read statements and pairs from an xlsx file with two sheets."""
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] openpyxl is required for xlsx files. Install: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True)

    # Read Statements sheet
    if "Statements" not in wb.sheetnames:
        print("[ERROR] xlsx must have a sheet named 'Statements'")
        sys.exit(1)
    ws = wb["Statements"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    statements = []
    for row in rows[1:]:
        if not row[0]:
            continue
        d = dict(zip(header, row))
        statements.append({
            "id": str(d.get("id", "")).strip(),
            "text": str(d.get("text_he", "")).strip(),
            "trait": str(d.get("trait", "")).strip(),
        })

    # Read Pairs sheet
    if "Pairs" not in wb.sheetnames:
        print("[ERROR] xlsx must have a sheet named 'Pairs'")
        sys.exit(1)
    ws = wb["Pairs"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    pairs = []
    for row in rows[1:]:
        if not row[0]:
            continue
        d = dict(zip(header, row))
        pairs.append({
            "id": str(d.get("id", "")).strip(),
            "statementA": str(d.get("statement_a", "")).strip(),
            "statementB": str(d.get("statement_b", "")).strip(),
        })

    return statements, pairs


def read_csv_files(statements_path, pairs_path):
    """Read statements and pairs from separate CSV files."""
    statements = []
    with open(statements_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            statements.append({
                "id": row["id"].strip(),
                "text": row["text_he"].strip(),
                "trait": row["trait"].strip(),
            })

    pairs = []
    with open(pairs_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pairs.append({
                "id": row["id"].strip(),
                "statementA": row["statement_a"].strip(),
                "statementB": row["statement_b"].strip(),
            })

    return statements, pairs


def validate(statements, pairs):
    """Validate data integrity. Returns (errors, warnings)."""
    errors = []
    warnings = []
    stmt_ids = set()
    stmt_traits = {}

    # Check statements
    for s in statements:
        if s["id"] in stmt_ids:
            errors.append(f"Duplicate statement ID: {s['id']}")
        stmt_ids.add(s["id"])

        if s["trait"] not in VALID_TRAITS:
            errors.append(f"Statement {s['id']}: invalid trait '{s['trait']}' (valid: {VALID_TRAITS})")
        else:
            stmt_traits[s["id"]] = s["trait"]

        if not s["text"]:
            errors.append(f"Statement {s['id']}: empty text")

    # Check pairs
    pair_ids = set()
    trait_counts = Counter()

    for p in pairs:
        if p["id"] in pair_ids:
            errors.append(f"Duplicate pair ID: {p['id']}")
        pair_ids.add(p["id"])

        if p["statementA"] not in stmt_ids:
            errors.append(f"Pair {p['id']}: statementA '{p['statementA']}' not found")
        if p["statementB"] not in stmt_ids:
            errors.append(f"Pair {p['id']}: statementB '{p['statementB']}' not found")

        trait_a = stmt_traits.get(p["statementA"])
        trait_b = stmt_traits.get(p["statementB"])

        if trait_a and trait_b and trait_a == trait_b:
            errors.append(f"Pair {p['id']}: both statements have same trait '{trait_a}'")

        if trait_a:
            trait_counts[trait_a] += 1
        if trait_b:
            trait_counts[trait_b] += 1

    # Check trait balance
    counts = [trait_counts.get(t, 0) for t in VALID_TRAITS]
    if len(set(counts)) > 1:
        warnings.append(f"Trait balance is uneven: {dict(trait_counts)}")
    else:
        print(f"[OK] Trait balance: {counts[0]} appearances each")

    return errors, warnings


def export_to_csv(data, output_dir):
    """Export questions.json data to CSV files for Google Sheet import."""
    os.makedirs(output_dir, exist_ok=True)

    with open(os.path.join(output_dir, "statements.csv"), "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["id", "text_he", "trait", "notes"])
        for s in data["statements"]:
            w.writerow([s["id"], s["text"], s["trait"], ""])

    with open(os.path.join(output_dir, "pairs.csv"), "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["id", "statement_a", "statement_b", "notes"])
        for p in data["pairs"]:
            w.writerow([p["id"], p["statementA"], p["statementB"], ""])

    print(f"[OK] Exported CSVs to {output_dir}")


def main():
    parser = argparse.ArgumentParser(description="Export Google Sheet to questions.json (or vice versa)")
    parser.add_argument("--xlsx", help="Path to xlsx file with Statements and Pairs sheets")
    parser.add_argument("--statements", help="Path to statements CSV file")
    parser.add_argument("--pairs", help="Path to pairs CSV file")
    parser.add_argument("--output", required=True, help="Output path for questions.json")
    parser.add_argument("--export-csv", help="Export current questions.json to CSV files in this directory")
    args = parser.parse_args()

    # Reverse mode: export JSON to CSV
    if args.export_csv:
        with open(args.output, "r", encoding="utf-8") as f:
            data = json.load(f)
        export_to_csv(data, args.export_csv)
        return

    if args.xlsx:
        statements, pairs = read_xlsx(args.xlsx)
    elif args.statements and args.pairs:
        statements, pairs = read_csv_files(args.statements, args.pairs)
    else:
        print("[ERROR] Provide either --xlsx or both --statements and --pairs")
        sys.exit(1)

    print(f"[OK] Loaded {len(statements)} statements, {len(pairs)} pairs")

    errors, warnings = validate(statements, pairs)

    for w in warnings:
        print(f"[WARN] {w}")
    for e in errors:
        print(f"[ERROR] {e}")

    if errors:
        print(f"\n{len(errors)} error(s) found. Fix before exporting.")
        sys.exit(1)

    # Trait summary
    trait_stmt_counts = Counter(s["trait"] for s in statements)
    print("\nStatements per trait:")
    for t in sorted(VALID_TRAITS):
        print(f"  {t}: {trait_stmt_counts.get(t, 0)}")
    print(f"\nTotal pairs: {len(pairs)}")

    output_data = {"statements": statements, "pairs": pairs}

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)

    print(f"\n[OK] Written to {args.output}")


if __name__ == "__main__":
    main()
